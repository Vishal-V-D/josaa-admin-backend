import os
import json
import base64
from io import BytesIO
from fastapi import APIRouter, HTTPException, Request, UploadFile, File, Query

from supabase import create_client, Client
from dotenv import load_dotenv
import uuid
import openpyxl
import csv
from io import StringIO
import json

# ---------- Supabase Setup ----------
load_dotenv()
SUPABASE_URL: str = os.environ.get("SUPABASE_URL") or "https://drdxhvqstjlxguyqpetq.supabase.co"
SUPABASE_KEY: str = os.environ.get("SUPABASE_KEY") or "eyJhbGciOiJIUzI1NiIsInR5cCI6IkpXVCJ9.eyJpc3MiOiJzdXBhYmFzZSIsInJlZiI6ImRyZHhodnFzdGpseGd1eXFwZXRxIiwicm9sZSI6ImFub24iLCJpYXQiOjE3NTQyOTI0NTksImV4cCI6MjA2OTg2ODQ1OX0.xd8YlNV8qV58-n1BG5jvcMGmtkH5dWUh92xzKR4JAnI"
supabase: Client = create_client(SUPABASE_URL, SUPABASE_KEY)

# ---------- Router Instance ----------
router = APIRouter(prefix="/api")

# ---------- Constants ----------
TYPE_TO_TABLE = {
    'IIT': 'IITs',
    'IIIT': 'IIITs',
    'NIT': 'NITs',
    'GFTI': 'GFTS'
}


TABLE_MAPPING = {
    "IIT": "IITs",
    "IIIT": "IIITs",
    "NIT": "NITs",
    "GFTI": "GFTS", # Mapped 'GFTI' to your specific table 'GFTS'
    "GFTS": "GFTS"  # Just in case
}

def process_csv_with_smart_fill(file_stream, filter_name=None):
    print(f"\n🔹 [CSV PROCESSOR] Starting CSV Processing... (Filter: '{filter_name}')")
    try:
        raw_data = file_stream.read().decode("utf-8-sig")
        stream = StringIO(raw_data)
        reader = csv.reader(stream)

        headers = next(reader, None)
        if not headers:
            return []

        headers = [h.strip() for h in headers if h.strip()]

        # Identify filter column
        target_col_idx = -1
        if filter_name:
            lower_headers = [h.lower() for h in headers]
            possible_names = ["institute name", "institute", "college name", "name", "college"]
            for name in possible_names:
                if name in lower_headers:
                    target_col_idx = lower_headers.index(name)
                    break

        data = []
        filtered_data = []
        previous_row_values = [""] * len(headers)

        for row_idx, row in enumerate(reader):

            if len(row) < len(headers):
                row += [""] * (len(headers) - len(row))

            processed_row = {}
            current_row_values = []

            for col_idx, value in enumerate(row[:len(headers)]):
                cleaned_value = value.strip()
                should_fill = (row_idx > 0) and (cleaned_value == "") and (col_idx < 3)

                final_value = previous_row_values[col_idx] if should_fill else cleaned_value
                processed_row[headers[col_idx]] = final_value
                current_row_values.append(final_value)

            data.append(processed_row)
            previous_row_values = current_row_values

            # Filtering Logic
            if filter_name:
                match = False
                if target_col_idx != -1:
                    if filter_name.lower() in str(current_row_values[target_col_idx]).lower():
                        match = True
                else:
                    for val in current_row_values:
                        if filter_name.lower() in str(val).lower():
                            match = True
                            break
                if match:
                    filtered_data.append(processed_row)

        if filter_name:
            return filtered_data if filtered_data else data

        return data

    except Exception as e:
        print("❌ CSV ERROR:", e)
        raise e


# ---------------------------------------
# EXCEL PROCESSOR
# ---------------------------------------
def process_excel_with_smart_fill(file_stream, filter_name=None):
    print(f"\n🔸 [EXCEL PROCESSOR] Starting Excel Processing... (Filter: '{filter_name}')")
    try:
        wb = openpyxl.load_workbook(file_stream, data_only=True)
        sheet = wb.active

        headers = []
        header_row_index = -1

        # Find headers in first 20 rows
        for i, row in enumerate(sheet.iter_rows(min_row=1, max_row=20, values_only=True)):
            valid_cells = [str(c).strip() for c in row if c]
            if valid_cells:
                header_row_index = i + 1
                headers = [str(c).strip() for c in row if c]
                break

        if not headers:
            return []

        target_col_idx = -1
        if filter_name:
            lower_headers = [h.lower() for h in headers]
            possible = ["institute name", "institute", "college name", "name", "college"]
            for name in possible:
                if name in lower_headers:
                    target_col_idx = lower_headers.index(name)
                    break

        data = []
        filtered_data = []
        previous_row_values = [""] * len(headers)

        # Process data rows
        for row_idx, row in enumerate(sheet.iter_rows(
                min_row=header_row_index + 1, values_only=True)):

            if all(cell is None or str(cell).strip() == "" for cell in row):
                continue

            row_list = list(row)
            if len(row_list) < len(headers):
                row_list += [None] * (len(headers) - len(row_list))

            processed_row = {}
            current_row_values = []

            for col_idx, value in enumerate(row_list[:len(headers)]):
                v = str(value).strip() if value else ""
                should_fill = (len(data) > 0) and (v == "") and (col_idx < 3)
                final_value = previous_row_values[col_idx] if should_fill else v

                processed_row[headers[col_idx]] = final_value
                current_row_values.append(final_value)

            data.append(processed_row)
            previous_row_values = current_row_values

            # Filtering
            if filter_name:
                match = False
                if target_col_idx != -1:
                    if filter_name.lower() in str(current_row_values[target_col_idx]).lower():
                        match = True
                else:
                    for val in current_row_values:
                        if filter_name.lower() in str(val).lower():
                            match = True
                            break
                if match:
                    filtered_data.append(processed_row)

        if filter_name:
            return filtered_data if filtered_data else data

        return data

    except Exception as e:
        print("❌ EXCEL ERROR:", e)
        raise e


# ---------------------------------------
# FASTAPI ENDPOINT
# ---------------------------------------
@router.post("/convert-csv")
async def convert_csv(
    file: UploadFile = File(...),
    filter_name: str = Query(None)
):
    try:
        filename = file.filename.lower()

        if filename.endswith(".csv"):
            json_data = process_csv_with_smart_fill(await file.read(), filter_name)

        elif filename.endswith(".xlsx") or filename.endswith(".xls"):
            json_data = process_excel_with_smart_fill(file.file, filter_name)

        else:
            return JSONResponse(
                {"error": "Unsupported file format. Use CSV or Excel."},
                status_code=400
            )

        return JSONResponse(content=json_data, status_code=200)

    except Exception as e:
        print("❌ API ERROR:", e)
        return JSONResponse({"error": "Failed to process file", "details": str(e)}, 500)




@router.post("/update-college-order")
def update_college_order():
    try:
        data = request.json
        category = data.get('category') 
        items = data.get('items')       

        # --- 1. LOGGING: Print what we received ---
        print("\n🔵 [DEBUG] Received /update-college-order request")
        print(f"   Category: {category}")
        print(f"   Items Count: {len(items) if items else 0}")
        # ------------------------------------------

        if not category or not items:
            print("❌ [ERROR] Missing 'category' or 'items' in payload")
            return jsonify({"error": "Missing data"}), 400

        # Get table name
        table_name = TABLE_MAPPING.get(category)
        if not table_name:
            print(f"❌ [ERROR] No table found for category: {category}")
            return jsonify({"error": "Invalid category"}), 400

        print(f"   Target Table: {table_name}")

        # --- 2. FIX: ACTUAL UPDATE LOOP ---
        updated_count = 0
        
        for item in items:
            record_id = item['id']
            new_order = item['sort_order']
            
            # Print specific item being updated (Optional: comment out if too spammy)
            # print(f"   -> Updating ID {record_id} to Order {new_order}")

            # Execute Supabase Update
            # We use the supabase client defined globally in your app
            response = supabase.table(table_name)\
                .update({"sort_order": new_order})\
                .eq("id", record_id)\
                .execute()

            # Verify if update happened
            if response.data:
                updated_count += 1
            else:
                print(f"⚠️ [WARN] ID {record_id} not found or not updated.")

        # --- 3. LOGGING: Final Result ---
        print(f"✅ [SUCCESS] Updated {updated_count} records in '{table_name}'\n")
        
        return jsonify({
            "message": f"Updated order for {table_name}",
            "updated_count": updated_count
        }), 200

    except Exception as e:
        print(f"❌ [CRITICAL ERROR] {str(e)}")
        return jsonify({"error": str(e)}), 500
# ---------- College Endpoints ----------
@router.get("/iit")
def get_iits():
    try:
        res = supabase.table("IITs").select("*").execute()
        return res.data
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"Failed to fetch IITs: {e}")

@router.get("/iiit")
def get_iiits():
    try:
        res = supabase.table("IIITs").select("*").execute()
        return res.data
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"Failed to fetch IIITs: {e}")

@router.get("/nit")
def get_nits():
    try:
        res = supabase.table("NITs").select("*").execute()
        return res.data
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"Failed to fetch NITs: {e}")

@router.get("/gfti")
def get_gfts():
    try:
        res = supabase.table("GFTS").select("*").execute()
        return res.data
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"Failed to fetch GFTs: {e}")

@router.get("/all")
def get_all_colleges():
    all_data = {}
    try:
        for table in ["IITs", "IIITs", "NITs", "GFTS"]:
            res = supabase.table(table).select("*").execute()
            all_data[table] = res.data
        return all_data
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"Failed to fetch all colleges: {e}")

@router.get("/college/{id}/{type}")
def get_college_by_id_and_type(id: str, type: str):
    try:
        table_name = TYPE_TO_TABLE.get(type.upper())
        if not table_name:
            raise HTTPException(status_code=400, detail="Invalid college type")
        
        colleges_res = supabase.table("colleges").select("data").eq("uuid", id).single().execute()
        if not colleges_res.data:
            raise HTTPException(status_code=404, detail="College not found")
        full_data = colleges_res.data.get("data")
        
        basic_res = supabase.table(table_name).select("data").eq("id", id).single().execute()
        if not basic_res.data:
            raise HTTPException(status_code=404, detail="College not found")
        basic_data = basic_res.data.get("data")
        
        return {"full_data": full_data, "basic_data": basic_data}
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"Internal server error: {e}")

@router.put("/college/{id}/{type}")
def update_college(id: str, type: str, request: Request):
    try:
        data = request.json()
        table_name = TYPE_TO_TABLE.get(type.upper())
        if not table_name:
            raise HTTPException(status_code=400, detail="Invalid college type")
        
        type_table_update_data = {"data": data.get("basic_data")}
        res_type_table = supabase.table(table_name).update(type_table_update_data).eq("id", id).execute()
        
        full_data = data.get("full_data")
        if full_data:
            colleges_table_update_data = {"college_name": data.get("college_name"), "data": full_data}
            res_colleges = supabase.table("colleges").update(colleges_table_update_data).eq("uuid", id).execute()
        
        if res_type_table.data and (not full_data or res_colleges.data):
            return {"message": "College updated successfully"}
        else:
            raise HTTPException(status_code=500, detail="Failed to update college")
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"Internal server error: {e}")

@router.delete("/college/{id}/{type}")
def delete_college(id: str, type: str):
    try:
        table_name = TYPE_TO_TABLE.get(type.upper())
        if not table_name:
            raise HTTPException(status_code=400, detail="Invalid college type")
        
        res_colleges = supabase.table("colleges").delete().eq("uuid", id).execute()
        res_type_table = supabase.table(table_name).delete().eq("id", id).execute()
        
        if res_colleges.data and res_type_table.data:
            return {"message": "College deleted successfully"}
        else:
            raise HTTPException(status_code=500, detail="Failed to delete college")
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"Internal server error: {e}")

@router.post("/add-college")
async def add_new_college(request: Request):
    try:
        # Read JSON body correctly in FastAPI
        data = await request.json()
        print("Received add-college request:")
        print(json.dumps(data, indent=2))

        # Extract fields
        college_name = data.get("college_name")
        college_type = data.get("type")
        full_data = data.get("full_data")
        basic_data = data.get("basic_data")

        # Validate required fields
        if not college_name or not college_type or not full_data or not basic_data:
            raise HTTPException(
                status_code=400,
                detail="Missing required data (college_name, type, full_data, basic_data)"
            )

        # Resolve table name
        table_name = TYPE_TO_TABLE.get(college_type.upper())
        if not table_name:
            raise HTTPException(
                status_code=400,
                detail=f"Invalid college type '{college_type}'"
            )

        # Generate unique ID
        new_uuid = str(uuid.uuid4())

        # Insert into main "colleges" table
        colleges_insert_data = {
            "uuid": new_uuid,
            "college_name": college_name,
            "data": full_data
        }
        res_colleges = supabase.table("colleges").insert(colleges_insert_data).execute()

        # Insert into type-specific table
        type_table_insert_data = {
            "id": new_uuid,
            "data": basic_data
        }
        res_type_table = supabase.table(table_name).insert(type_table_insert_data).execute()

        # Validate insertion success
        if res_colleges.data and res_type_table.data:
            print("College added successfully!")
            return {
                "message": "College added successfully",
                "uuid": new_uuid
            }

        raise HTTPException(
            status_code=500,
            detail="Supabase failed to insert into one or both tables"
        )

    except Exception as e:
        print("🔥 ERROR in /add-college:", str(e))
        raise HTTPException(
            status_code=500,
            detail={"error": "Internal server error", "details": str(e)}
        )


@router.post("/upload-image")
def upload_image(request: Request):
    try:
        data = request.json()
        image_data = data.get("image_data")
        file_name = data.get("file_name")
        
        if not image_data or not file_name:
            raise HTTPException(status_code=400, detail="Invalid request body")
            
        image_bytes = base64.b64decode(image_data)
        unique_file_name = f"{uuid.uuid4()}_{file_name}"
        
        res = supabase.storage.from_("images").upload(unique_file_name, image_bytes)
        
        if res:
            return {"message": "Image uploaded successfully", "file_path": res.path}
        else:
            raise HTTPException(status_code=500, detail="Failed to upload image")
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"Internal server error: {e}")

@router.get("/college-type/{college_type}")
def get_colleges_by_type(college_type: str):
    try:
        res = supabase.table("Colleges").select("*").eq("type", college_type).execute()
        return res.data
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"Failed to fetch colleges of type '{college_type}': {e}")

# ---------- Announcements Endpoints ----------
@router.get("/announcements")
def get_announcements():
    try:
        res = supabase.table("announcements").select("*").execute()
        return res.data
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"Internal server error: {e}")

@router.post("/announcements")
def add_announcement(request: Request):
    try:
        data = request.json()
        title = data.get("title")
        data_json = data.get("data_json")
        
        if not title or not data_json:
            raise HTTPException(status_code=400, detail="Title and data_json are required")
            
        insert_data = {"title": title, "data_json": json.dumps(data_json), "id": str(uuid.uuid4())}
        res = supabase.table("announcements").insert(insert_data).execute()
        
        if res.data:
            return {"message": "Announcement added successfully"}
        else:
            raise HTTPException(status_code=500, detail="Failed to add announcement")
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"Internal server error: {e}")

@router.put("/announcements/{announcement_id}")
def update_announcement(announcement_id: str, request: Request):
    try:
        data = request.json()
        title = data.get("title")
        data_json = data.get("data_json")
        
        if not title or not data_json:
            raise HTTPException(status_code=400, detail="Title and data_json are required")
            
        update_data = {"title": title, "data_json": json.dumps(data_json)}
        res = supabase.table("announcements").update(update_data).eq("id", announcement_id).execute()
        
        if res.data:
            return {"message": "Announcement updated successfully"}
        else:
            raise HTTPException(status_code=500, detail="Failed to update announcement")
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"Internal server error: {e}")

@router.delete("/announcements/{announcement_id}")
def delete_announcement(announcement_id: str):
    try:
        res = supabase.table("announcements").delete().eq("id", announcement_id).execute()
        if res.data:
            return {"message": "Announcement deleted successfully"}
        else:
            raise HTTPException(status_code=500, detail="Failed to delete announcement")
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"Internal server error: {e}")
